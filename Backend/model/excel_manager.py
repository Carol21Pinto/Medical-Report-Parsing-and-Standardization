import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_FILE_PATH = "lab_results.xlsx"

def initialize_excel():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE_PATH):
        df = pd.DataFrame(columns=[
            'Patient_Name',
            'Age',
            'Sex',
            'UHID',
            'Episode',
            'Ref_Doctor',
            'Test_Name',
            'Test_Value',
            'Unit',
            'Reference_Range',
            'Status',
            'Bill_No',
            'Facility',
            'Sample_No',
            'Collection_Date',
            'Report_Date',
            'Extraction_Date'
        ])
        df.to_excel(EXCEL_FILE_PATH, index=False, sheet_name='Lab Results')
        format_excel_file()
        print(f"✓ Created new Excel file: {EXCEL_FILE_PATH}")
    return EXCEL_FILE_PATH


def format_excel_file():
    """Apply formatting to Excel file"""
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(EXCEL_FILE_PATH)
    except Exception as e:
        print(f"Warning: Could not format Excel file: {str(e)}")


def append_lab_results_to_excel(extracted_data):
    """
    Append lab test results from extracted data to Excel file
    
    Args:
        extracted_data: Dictionary containing patient info and lab tests
    
    Returns:
        Number of rows added
    """
    initialize_excel()
    
    patient_info = extracted_data.get('patient_information', {})
    order_info = extracted_data.get('order_information', {})
    lab_tests = extracted_data.get('lab_tests', [])
    
    if not lab_tests:
        print("⚠ No lab tests found to add to Excel")
        return 0
    
    # Prepare rows for Excel
    new_rows = []
    extraction_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    for test in lab_tests:
        row = {
            'Patient_Name': patient_info.get('patient_name', 'N/A'),
            'Age': patient_info.get('age', 'N/A'),
            'Sex': patient_info.get('sex', 'N/A'),
            'UHID': patient_info.get('uhid', 'N/A'),
            'Episode': patient_info.get('episode', 'N/A'),
            'Ref_Doctor': patient_info.get('ref_doctor', 'N/A'),
            'Test_Name': test.get('test_name', 'N/A'),
            'Test_Value': test.get('value', 'N/A'),
            'Unit': test.get('unit', 'N/A'),
            'Reference_Range': test.get('reference_range', 'N/A'),
            'Status': test.get('status', 'N/A'),
            'Bill_No': order_info.get('bill_no', 'N/A'),
            'Facility': order_info.get('facility', 'N/A'),
            'Sample_No': order_info.get('sample_no', 'N/A'),
            'Collection_Date': order_info.get('collection_date', 'N/A'),
            'Report_Date': order_info.get('report_date', 'N/A'),
            'Extraction_Date': extraction_date
        }
        new_rows.append(row)
    
    # Read existing data and append new rows
    try:
        existing_df = pd.read_excel(EXCEL_FILE_PATH)
        new_df = pd.DataFrame(new_rows)
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        combined_df.to_excel(EXCEL_FILE_PATH, index=False, sheet_name='Lab Results')
        
        # Apply formatting
        format_excel_file()
        
        print(f"✓ Added {len(new_rows)} lab test records to Excel")
        return len(new_rows)
        
    except Exception as e:
        print(f"✗ Error appending to Excel: {str(e)}")
        return 0


def get_excel_stats():
    """Get statistics about the Excel file"""
    if not os.path.exists(EXCEL_FILE_PATH):
        return {
            'exists': False,
            'total_records': 0,
            'unique_patients': 0
        }
    
    try:
        df = pd.read_excel(EXCEL_FILE_PATH)
        return {
            'exists': True,
            'total_records': len(df),
            'unique_patients': df['Patient_Name'].nunique() if 'Patient_Name' in df.columns else 0,
            'file_path': EXCEL_FILE_PATH
        }
    except Exception as e:
        return {
            'exists': True,
            'error': str(e)
        }
